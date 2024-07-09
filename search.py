from glob import glob
import openpyxl
from os import path,mkdir
from time import time

start_time = time()

OUTPUT_PATH = './output/'

if not path.exists(OUTPUT_PATH):
    mkdir(OUTPUT_PATH)

files = glob('./files/*')

for index,single_file in enumerate(files):
    print(str(index) + ' ' + single_file)

user_input = str(input('enter coloum index: ')).split(',')

for file_index,file in enumerate(files):

    file_name = file.split('\\')[1].split('.')[0]
    
    excel = openpyxl.load_workbook(file)

    for sheet in excel:
        active_excel = excel[sheet.title]
        null_counter = 0

        for row_index,row in enumerate(active_excel.iter_rows(values_only=True)):
            
            current_target = str(row[int(user_input[file_index])]).replace(' ','').replace('-','').replace('_','').replace(':','').replace('/','').strip()

            if null_counter >= 200:
                break

            if row[int(user_input[file_index])] == None:
                null_counter += 1

            if row[int(user_input[file_index])] == None or row[int(user_input[file_index])] == '':
                continue

            null_counter = 0

            print(str(row_index) + ': ' + current_target)
            
            for search_sheet in excel:
                
                search_active_excel = excel[search_sheet.title]
                search_null_counter = 0

                for search_row_index,search_row in enumerate(search_active_excel.iter_rows(values_only=True)):

                    search_value = str(search_row[int(user_input[file_index])]).replace(' ','').replace('-','').replace('_','').replace(':','').replace('/','').strip()

                    if search_null_counter >= 200:
                        break

                    if search_row[int(user_input[file_index])] == None:
                        search_null_counter += 1

                    if search_row[int(user_input[file_index])] == None or search_row[int(user_input[file_index])] == '':
                        continue

                    search_null_counter = 0

                    if current_target == search_value or search_value.find(current_target) > -1:
                        
                        output_file = open(OUTPUT_PATH + file_name + '.txt','a',encoding='utf-8')
                        output_file.write(f"search word: {current_target} | sheet: {search_sheet.title} | row: {search_row_index + 1} | finded_Value: {search_value} \n\n")

                        output_file.close()

print(time() - start_time)