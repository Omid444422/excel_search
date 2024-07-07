from glob import glob
import openpyxl

user_search_input = str(input('enter search word: ')).strip()

user_search_input = user_search_input.replace(' ','').replace('-','').replace('_','').replace(':','').replace('/','')


# find all excels
files = glob('./files/*')

founded_result = list()

for file in files:

    current_excel_file = openpyxl.load_workbook(file)
    
    file_name = file.split('\\')[1]

    sheets = list()

    # add every sheets name to sheets
    for sheet in current_excel_file:
        sheets.append(sheet.title)

        # loop on every exist sheet
    for single_sheet in sheets:
        current_excel_sheet = current_excel_file[single_sheet]
        null_counter = 0

        is_ended = False

        for row_id,row in enumerate(current_excel_sheet.iter_rows(0,current_excel_sheet.max_row,values_only=True)):

            for single_value in row:
                if is_ended:
                    break
                
                current_value = str(single_value).replace(' ','').replace('-','').replace('_','').replace(':','').replace('/','')

                if (str(current_value)).find(user_search_input) > -1 or current_value == user_search_input:

                    is_exist = False

                    for result in founded_result:
                        if result['row'] == row_id +1:
                            is_exist = True
                            continue
                        else:
                            is_exist = False

                    if not is_exist:
                        founded_result.append({'file_name':file_name,'sheet_name':single_sheet,'row':row_id + 1,'founded_value':str(single_value)})

                if single_value == None:
                    null_counter += 1
                else:
                    null_counter = 0

                if single_value == None and null_counter >= 200:
                    is_ended = True
                    break

            if is_ended == True:
                break

for single_found in founded_result:
    print(f'file name: {single_found['file_name']} sheet: {single_found['sheet_name']} row: {single_found['row']} value: {single_found['founded_value']} \n')

exit()
